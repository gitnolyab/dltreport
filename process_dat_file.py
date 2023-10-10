import openpyxl
import random  # THIS IS FOR RANDOM ALPHA NUMERIC AT THE END
import string  # THIS IS FOR RANDOM ALPHA NUMERIC AT THE END
import streamlit as st
# import os  # use to move dat file data
# import shutil  # use to move dat file data
import time


# def process_dat(custom_file_path, user_input_SheetName):
def process_dat(custom_file_path, user_input_directory):
    # KANI ANG RUNNING SA DAT FILE PRINTING BQY, ANG KULANG NALANG IS DAT FILE CLEANING,PROCESSING

    vH1702Q = 'H1702Q'
    vHASWT = 'HSAWT'
    vTIN = '004357072'
    vPAYEES = '"MEGAFARM INTEGRATED AGRO LIVESTOCK FARM INC"'
    vBranch_cdo = '0000'
    vFormCode = '1702Q'
    vPeriod = ''
    vDSAWT = 'DSAWT'
    vD1702Q = 'D1702Q'
    vSEQ_NO = 1
    vDTIN = ''
    svTin3 = ''
    svTin = ''
    svbranch = ''
    # vPrint_Code = 'GO'  # use to give signal to print

    svCorpName = ''  # None  # save corp name to check current company
    svIndivName = ''  # None  # save individual       (save only if ? ask bqy )
    svAtc_Code = ''  # None  # save atc_code to check the current ATC_CODE
    svAmtOfPayment = 0.00  # saved amount of income payment
    svAmtofTWHeld = 0.00  # saved amount of tax withheld

    cell_C_value = ''  # CORP
    cell_D_value = ''  # INDI

    acc_cell_G_value = 0.00  # accumulator initialize
    acc_cell_I_value = 0.00  # accumulator initialize
    acc_Total_cell_G_value = 0.00  # grand total column G accumulator
    acc_Total_cell_I_value = 0.00  # grand total column I accumulator

    cell_G_value = 0.00
    cell_H_value = 0.00
    cell_I_value = 0.00

    # Define a dictionary to map month names to numerical values
    month_mapping = {
        'JANUARY': '01',
        'FEBRUARY': '02',
        'MARCH': '03',
        'APRIL': '04',
        'MAY': '05',
        'JUNE': '06',
        'JULY': '07',
        'AUGUST': '08',
        'SEPTEMBER': '09',
        'OCTOBER': '10',
        'NOVEMBER': '11',
        'DECEMBER': '12'
    }

    # workbook = openpyxl.load_workbook("SAWT1.xlsx")           # from previous running code
    # sheet1 = workbook['IST']                                  # from previous running code
    workbook = openpyxl.load_workbook(custom_file_path)
    sheet1 = workbook.active

    # Get the value from cell A3
    A3 = sheet1['A3'].value

    # Initialize variables to store month and year
    month = None
    year = None

    # Check if A3 is not empty and contains a valid date format
    if A3:
        # Split the string by ', ' and select the second part

        parts = A3.split(', ')
        # parts = ('FOR THE MONTH OF MARCH', '2023')

        if len(parts) >= 2:
            # Extract the month and year from the string
            try:
                year = parts[1]  # ok na
                # year = 2023

                # Extract the month and year by splitting the first part by space
                month_year = parts[0].split(' ')  # split by space
                # month_year = ("FOR", "THE", "MONTH", "OF", "MARCH")

                month = month_year[-1].strip(',')
                # month = MARCH
                month_name = month.upper()  # Capitalized lang para sure
                # month_name = MARCH

                # check kung ang month is naa sa dictionary, kung naa replace with numeric
                if month_name in month_mapping:  # e search niya sa month_mapping dictionary
                    numerical_month = month_mapping[month_name]  # kung makita replace to number
                    # numerical_month = '03'

                    vPeriod = f"{numerical_month}{year}"
                    # vPeriod = '032023'        # ok na ang datfile creation...

            except ValueError:
                print("Invalid date format in cell A3.")
        else:
            print("Sheet A3 does not contain the expected date format.")
    else:
        print("Sheet A3 is empty.")

    # Start from row 16 (assuming row 15 is the header)
    current_row = 16
    # Determine the maximum row in the Excel sheet
    max_row = sheet1.max_row

    # Set the maximum column number you want to process
    max_column_number = 9

    # filename creation from variable, tin, vBranch_cdo, period and code
    dat_filename = f"{vTIN}{vBranch_cdo}{vPeriod}{vFormCode}"

    fname = ''
    lname = ''
    mname = ''

    # this is just to fill "","","" in the first row
    if fname in (None, 'None', ''):
        fname = '""'
    if lname in (None, 'None', ''):
        lname = '""'
    if mname in (None, 'None', ''):
        mname = '""'

    # <<<<<<<<<<<< -------STARTS OPENING OUTPUT FILE ------------
    # with (open(f"{user_input_directory} \ {dat_filename}.dat", "w") as output_file):

    with (open(f"{dat_filename}.dat", "w") as output_file):
        # first line to be printed in the output file..
        vFirstLine = f"{vHASWT},{vH1702Q},{vTIN},{vBranch_cdo},{vPAYEES},{fname},{lname},{mname},{vPeriod},098"
        output_file.write(vFirstLine + "\n")

        # FIRST READING & ACCESS OF VALUE FROM (EXCEL FILE)
        cell_A_value, cell_B_value, cell_C_value, cell_D_value, cell_E_value, cell_F_value, cell_G_value, \
            cell_H_value, cell_I_value = (str(sheet1.cell(row=current_row, column=i).value).strip() for i in
                                          range(1, 10))
        # READY FOR NEXT RECORD
        current_row += 1

        # FIRST SAVING OF DATA
        # svSeq_No = cell_A_value
        svTin = cell_B_value
        svCorpName = cell_C_value  # Corporation
        svIndivName = cell_D_value  # Individual
        svAtc_Code = cell_E_value
        svAmtOfPayment = round(float(cell_G_value), 2)
        svTaxRate = round(float(cell_H_value), 2)
        svAmtofTWHeld = round(float(cell_I_value), 2)

        # FIRST ACCUMULATION AT FIRST SAVING
        acc_cell_G_value += round(float(svAmtOfPayment), 2)
        acc_cell_I_value += round(float(svAmtofTWHeld), 2)
        acc_Total_cell_G_value += round(float(svAmtOfPayment), 2)
        acc_Total_cell_I_value += round(float(svAmtofTWHeld), 2)

        # MAIN LOOP
        while current_row <= max_row:

            # NEXT READ FILE (EXCEL DATA) (CELL_F_VALUE is note include)
            cell_A_value, cell_B_value, cell_C_value, cell_D_value, cell_E_value, cell_F_value, cell_G_value, \
                cell_H_value, cell_I_value = (str(sheet1.cell(row=current_row, column=i).value).strip() for i in
                                              range(1, 10))

            # Data validation st    cell_A_value is EMPTY ---<<<---
            if cell_A_value is None:
                break

            # Check if 'Grand' is in cell_A_value (case-sensitive)
            if 'Grand' in cell_A_value:
                break

            if '---' in cell_I_value:
                break

            # CHECK TIN == svTin
            if cell_B_value == svTin:
                # svTin = cell_B_value      # From Save routine

                # CHECK IF ATC_CODE == SAVE ATC_CODE (IF NOT EQUAL CALL PRINT RTN)   ---
                if cell_E_value == svAtc_Code:

                    # ACCUMULATE TAX AMOUNT & GRAND TOTAL
                    if cell_G_value != 'None':
                        acc_cell_G_value += round(float(cell_G_value), 2)  # OK NA, nag accumulate na
                        acc_Total_cell_G_value += round(float(cell_G_value), 2)  # grand total column I acc

                    if cell_I_value != 'None':
                        acc_cell_I_value += round(float(cell_I_value), 2)  # OK NA, nag accumulate na
                        acc_Total_cell_I_value += round(float(cell_I_value), 2)  # grand total column I accumulator

                    # INCREMENT RECORD READ CTR
                    current_row += 1  # READY FOR NEXT RECORD

                    # ENDIF SA ACCUMULATE TAXES

                else:

                    # PRINT ROUTINE --- FROM CHECK IF ATC_CODE == SAVE ATC_CODE

                    # CHECK IF CORPORATION/INDIVIDUAL NAME IS EMPTY ---
                    # corporation
                    if cell_C_value in (None, 'None', ''):
                        cell_C_value = ''  # two commas
                        # svCorpName = ''  # cell_C_value

                    # individual
                    # Check and replace four commas if cell_D_value has no value or is 'None'
                    if cell_D_value in (None, 'None', ''):
                        cell_D_value = ',,'  # four commas
                        # svIndivName = ',,'  # cell_D_value

                    # TIN & BRANCH CODE EXTRACTION
                    cell_b2 = svTin.split('-')  # cell_b2: ['006', '430', '305', '0002']
                    branch_1 = cell_b2[-1]  # "0002" is the branch      # Extract the last group
                    tin_3 = ''.join(cell_b2[:3])  # "0064330305 is the TIN
                    svTin3 = tin_3
                    svbranch = branch_1

                    vR2_acc_G = round(acc_cell_G_value, 2)
                    acc_cell_G_value = vR2_acc_G
                    vR2_acc_I = round(acc_cell_I_value, 2)
                    acc_cell_I_value = vR2_acc_I

                    # MOVE SAVED DATA TO VARIABLES READY FOR PRINTING
                    to_be_join_row_data = [vDSAWT,
                                           vD1702Q,
                                           vSEQ_NO,
                                           svTin3,
                                           svbranch,
                                           svCorpName,
                                           svIndivName,
                                           vPeriod,
                                           svAtc_Code,
                                           svTaxRate,
                                           acc_cell_G_value,
                                           acc_cell_I_value,
                                           vTIN,  # MAIN
                                           vBranch_cdo
                                           ]

                    # JOIN DATA VARIABLE READY FOR PRINTING
                    row_data_str = ",".join(map(str, to_be_join_row_data))

                    # PRINT NEW ROW TO DAT FILE
                    output_file.write(row_data_str + "\n")

                    # ZERO-OUT ROW ACCUMULATORS
                    acc_cell_G_value = 0.00  # row accumulator initialize
                    acc_cell_I_value = 0.00  # row accumulator initialize

                    # SAVE CURRENT READ DATA (JUST READ)
                    # NEXT SAVING OF DATA
                    svTin = cell_B_value
                    svCorpName = cell_C_value  # Corporation
                    svIndivName = cell_D_value  # Individual
                    svAtc_Code = cell_E_value
                    svAmtOfPayment = round(float(cell_G_value), 2)
                    svTaxRate = round(float(cell_H_value), 2)
                    svAmtofTWHeld = round(float(cell_I_value), 2)

                    # ACCUMULATE CURRENT READ DATA
                    # ACCUMULATE TAX AMOUNT & GRAND TOTAL
                    if cell_G_value != 'None':
                        acc_cell_G_value += round(float(cell_G_value), 2)  # OK NA, nag accumulate na
                        acc_Total_cell_G_value += round(float(cell_G_value), 2)  # grand total column I acc

                    if cell_I_value != 'None':
                        acc_cell_I_value += round(float(cell_I_value), 2)  # OK NA, nag accumulate na
                        acc_Total_cell_I_value += round(float(cell_I_value), 2)  # grand total column I accumulator

                    # INCREMENT ROW CTR/SEQ_CTR AFTER PRINTING
                    vSEQ_NO += 1

                    # INITIALIZE NEXT ROW
                    current_row += 1

                # ENDIF SA ATC_CODE CHECKING

            else:
                # PRINT ROUTINE --- FROM TIN CHECKING

                # IF SV_CORP == 'NONE'
                # IF SV_INDI == 'NONE'

                # CHECK IF CORPORATION/INDIVIDUAL NAME IS EMPTY ---
                # corporation
                if cell_C_value in (None, 'None', ''):
                    cell_C_value = ''  # two commas
                    # svCorpName = ''  # cell_C_value

                # individual
                # Check and replace four commas if cell_D_value has no value or is 'None'
                if cell_D_value in (None, 'None', ''):
                    cell_D_value = ',,'  # four commas
                    # svIndivName = ',,'  # cell_D_value

                # TIN & BRANCH CODE EXTRACTION
                cell_b2 = svTin.split('-')  # cell_b2: ['006', '430', '305', '0002']
                branch_1 = cell_b2[-1]  # "0002" is the branch      # Extract the last group
                tin_3 = ''.join(cell_b2[:3])  # "0064330305 is the TIN
                svTin3 = tin_3
                svbranch = branch_1

                vR2_acc_G = round(acc_cell_G_value, 2)
                acc_cell_G_value = vR2_acc_G
                vR2_acc_I = round(acc_cell_I_value, 2)
                acc_cell_I_value = vR2_acc_I

                # MOVE SAVED DATA TO VARIABLES READY FOR PRINTING
                to_be_join_row_data = [vDSAWT,
                                       vD1702Q,
                                       vSEQ_NO,
                                       svTin3,
                                       svbranch,
                                       svCorpName,
                                       svIndivName,
                                       vPeriod,
                                       svAtc_Code,
                                       svTaxRate,
                                       acc_cell_G_value,
                                       acc_cell_I_value,
                                       vTIN,  # MAIN
                                       vBranch_cdo
                                       ]

                # JOIN DATA VARIABLE READY FOR PRINTING
                row_data_str = ",".join(map(str, to_be_join_row_data))

                # PRINT NEW ROW TO DAT FILE
                output_file.write(row_data_str + "\n")

                # ZERO-OUT ROW ACCUMULATORS
                acc_cell_G_value = 0.00  # row accumulator initialize
                acc_cell_I_value = 0.00  # row accumulator initialize

                # SAVE CURRENT READ DATA (JUST READ)
                # NEXT SAVING OF DATA
                svTin = cell_B_value
                svCorpName = cell_C_value  # Corporation
                svIndivName = cell_D_value  # Individual
                svAtc_Code = cell_E_value
                svAmtOfPayment = round(float(cell_G_value), 2)
                svTaxRate = round(float(cell_H_value), 2)
                svAmtofTWHeld = round(float(cell_I_value), 2)

                # ACCUMULATE CURRENT READ DATA
                # ACCUMULATE TAX AMOUNT & GRAND TOTAL
                if cell_G_value != 'None':
                    acc_cell_G_value += round(float(cell_G_value), 2)  # OK NA, nag accumulate na
                    acc_Total_cell_G_value += round(float(cell_G_value), 2)  # grand total column I acc

                if cell_I_value != 'None':
                    acc_cell_I_value += round(float(cell_I_value), 2)  # OK NA, nag accumulate na
                    acc_Total_cell_I_value += round(float(cell_I_value), 2)  # grand total column I accumulator

                # INCREMENT ROW CTR/SEQ_CTR AFTER PRINTING
                vSEQ_NO += 1

                # INITIALIZE NEXT ROW
                current_row += 1

        # ENDIF OF TIN CHECKING

        # END WHILE LOOP HERE ------------------------------

        # PRINT LAST READ DATA W/SAVED

        # TIN & BRANCH CODE EXTRACTION
        cell_b2 = svTin.split('-')  # cell_b2: ['006', '430', '305', '0002']
        branch_1 = cell_b2[-1]  # "0002" is the branch      # Extract the last group
        tin_3 = ''.join(cell_b2[:3])  # "0064330305 is the TIN    # Merge the first three groups without a hyphen
        svTin3 = tin_3
        svbranch = branch_1

        vR2_acc_G = round(acc_cell_G_value, 2)
        acc_cell_G_value = vR2_acc_G
        vR2_acc_I = round(acc_cell_I_value, 2)
        acc_cell_I_value = vR2_acc_I

        # MOVE SAVED DATA TO VARIABLES READY FOR PRINTING
        to_be_join_row_data = [vDSAWT,
                               vD1702Q,
                               vSEQ_NO,
                               svTin3,
                               svbranch,
                               svCorpName,
                               svIndivName,
                               vPeriod,
                               svAtc_Code,
                               svTaxRate,
                               acc_cell_G_value,
                               acc_cell_I_value,
                               vTIN,  # MAIN
                               vBranch_cdo
                               ]

        # JOIN DATA VARIABLE READY FOR PRINTING
        row_data_str = ",".join(map(str, to_be_join_row_data))

        # PRINT LAST RECORD TO DAT FILE
        output_file.write(row_data_str + "\n")

        # ADDING 1 SPACE FOR TOTAL IN PRINTING
        total_G_acc = f' {round(acc_Total_cell_G_value, 2)}'
        acc_Total_cell_G_value = total_G_acc

        total_I_acc = f' {round(acc_Total_cell_I_value, 2)}'
        acc_Total_cell_I_value = total_I_acc

        # PRINT GRAND TOTAL ACCUMULATORS HERE...
        lastdata = ['CSAWT',
                    'C1702Q',
                    vTIN,
                    vBranch_cdo,
                    vPeriod,
                    acc_Total_cell_G_value,
                    acc_Total_cell_I_value
                    ]

        # Join last data variable
        row_data_str = ",".join(map(str, lastdata))

        # print to last Data to DAT file
        output_file.write(row_data_str + "\n")

        st.warning('Processing DAT file...')
        st.success('DAT FILE GENERATED PLEASE CHECK YOU DIRECTORY DRIVE ...')

        # GENERATE RANDOM ALPHA NUMERIC AT THE END OF DAT FILE
        # Define the length of the alphanumeric string
        length = 21  # You can change this to the desired length
        # Define the characters to use in the alphanumeric string
        characters = string.ascii_uppercase + string.digits  # Uppercase letters and digits
        # Generate a random alphanumeric string
        random_string = 'V' + ''.join(random.choice(characters) for _ in range(length - 1))
        # print(f"Random alphanumeric string: {random_string}")
        output_file.write(random_string)  # MORAG AUTO GENERATED NI

        # move data to user_input_directory
        # os.chdir('mkdir ')

        # source = output_file
        # destination = f"D:\\bir\\sawt\\{output_file}"
        # shutil.move(f"{source},{destination}")
        #
        # # try:
        #     os.replace(output_file, destination)
        #     print(source + " was moved")
        # except FileNotFoundError:
        #     print('not moved')

        import os
        # import shutil
        #
        # # Define the source and destination paths
        # source = f"{output_file}"
        # print("source :", source)
        # # source = "path_to_source_file"  # Replace with the actual path to your source file
        # destination = f"D:\\bir\\sawt\\{output_file}"  # Replace with the desired destination directory
        #
        # # Move the file
        # shutil.move(source, destination)

# --------------------- END OF OPENING OUTPUT FILE >>>>>>>>>>>>>>>>>>>>>>>>>>>
